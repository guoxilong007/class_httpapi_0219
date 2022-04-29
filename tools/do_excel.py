import pprint

from openpyxl import load_workbook
from tools.do_config import Do_Config
from tools import get_path
from tools import get_data
from tools.do_regal import Do_Regal



class Do_Excel():

    # 读取excel内测试用户数据
    def do_excel(self, file_name):
        # 获取到配置文件内的excel表单名
        test_case = eval(Do_Config.do_config(get_path.conf_path, "CASE", "case"))
        # 获取到excel内init表单内的手机号
        timing_phone = getattr(get_data.Get_Data, "timing_phone")
        # 创建一个空列表来收集测试用例数据
        test_data = []
        # 创建excel连接方式
        wb = load_workbook(file_name)
        # 将sheet名称遍历出来
        for sheet_name in test_case:
            # 将sheet_name用遍历的方式进行传递
            sheet = wb[sheet_name]
            """
            根据配置文件的value来读取测试用例；
            value等于all时，则读取该sheet内的全部测试用例数据
            value等于列表数据时，则根据列表内的数字读取对应case_id的测试用例数据"""
            if test_case[sheet_name] == "all":
                # 读取全部的测试用例数据：
                for i in range(2, sheet.max_row+1):
                    test_dict = {}
                    test_dict["case_id"] = sheet.cell(i, 1).value
                    test_dict["module"] = sheet.cell(i, 2).value
                    test_dict["title"] = sheet.cell(i, 3).value
                    test_dict["url"] = sheet.cell(i, 4).value
                    if sheet.cell(i, 5).value.find("${tel_1}") != -1:
                        test_dict["data"] = sheet.cell(i, 5).value.replace("${tel_1}", str(timing_phone+1))
                    elif sheet.cell(i, 5).value.find("${tel}") != -1:
                        test_dict["data"] = sheet.cell(i, 5).value.replace("${tel}", str(timing_phone+2))
                    else:
                        test_dict["data"] = sheet.cell(i, 5).value
                    if sheet.cell(i, 6).value != None:
                        # 使用正则表达式方法对excel表内的数据进行读取
                        test_dict["check_sql"] = Do_Regal.do_regal(sheet.cell(i, 6).value)
                    else:
                        test_dict["check_sql"] = None
                    test_dict["method"] = sheet.cell(i, 7).value
                    test_dict["expected"] = sheet.cell(i, 8).value
                    # 把遍历出来的sheet_name对应起来
                    test_dict["sheet_name"] = sheet_name
                    test_data.append(test_dict)
                    self.new_phone(get_path.case_path, "init", str(timing_phone+3))
            else:
                # 遍历配置文件内的列表数据，列表内的数值对应这用例ID，使用case_id变量进行接收；
                for case_id in test_case[sheet_name]:
                    test_dict = {}
                    # 根据行号读取测试用例时，case_id需要+1，因为excel内的表单存在title
                    test_dict["case_id"] = sheet.cell(case_id+1, 1).value
                    test_dict["module"] = sheet.cell(case_id+1, 2).value
                    test_dict["title"] = sheet.cell(case_id+1, 3).value
                    test_dict["url"] = sheet.cell(case_id+1, 4).value
                    if sheet.cell(case_id+1, 5).value.find("${tel_1}") != -1:
                        test_dict["data"] = sheet.cell(case_id+1, 5).value.replace("${tel_1}", str(timing_phone+1))
                    elif sheet.cell(case_id+1, 5).value.find("${tel}") != -1:
                        test_dict["data"] = sheet.cell(case_id+1, 5).value.replace("${tel}", str(timing_phone+2))
                    else:
                        test_dict["data"] = sheet.cell(case_id+1, 5).value
                    if sheet.cell(case_id+1, 6).value != None:
                        test_dict["check_sql"] = Do_Regal.do_regal(str(sheet.cell(case_id+1, 6).value))
                    else:
                        test_dict["check_sql"] = None
                    test_dict["method"] = sheet.cell(case_id+1, 7).value
                    test_dict["expected"] = sheet.cell(case_id+1, 8).value
                    test_dict["sheet_name"] = sheet_name
                    test_data.append(test_dict)
                    self.new_phone(get_path.case_path, "init", str(timing_phone+3))

        return test_data

    # 将数据写回至excel方法
    def result_back(self, file_name, sheet_name, row, col, test_result):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(row, col).value = test_result
        wb.save(file_name)


    # 每次请求时更新手机号码
    def new_phone(self, file_name, sheet_name, new_phone):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(2, 2).value = new_phone
        wb.save(file_name)



if __name__ == '__main__':
    res_1 = Do_Excel().do_excel(get_path.case_path)
    pprint.pprint(res_1)






