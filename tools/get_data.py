import pandas as pd
from tools import get_path
from openpyxl import load_workbook


class Get_Data():

    COOKIES = None
    # 使用pandas从excel内获取到手机号
    timing_phone = pd.read_excel(get_path.case_path, sheet_name="init").iloc[0,1]
    # 储存userID的数据
    userID = None

    # 使用openpyxl从excel内获取到手机号
    wb = load_workbook(get_path.case_path)
    sheet = wb["init"]
    timing_phone2 = sheet.cell(2,2).value



if __name__ == '__main__':
    res = Get_Data().timing_phone
    res2 = Get_Data().timing_phone2
    print(type(res))
    print(res2)